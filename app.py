import io, math, os, base64, tempfile
from flask import Flask, request, send_file
from flask_cors import CORS
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage

app = Flask(__name__)
CORS(app)

LOGO_B64 = "/9j/4AAQSkZJRgABAQAAAQABAAD/4gHYSUNDX1BST0ZJTEUAAQEAAAHIAAAAAAQwAABtbnRyUkdCIFhZWiAH4AABAAEAAAAAAABhY3NwAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAQAA9tYAAQAAAADTLQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAlkZXNjAAAA8AAAACRyWFlaAAABFAAAABRnWFlaAAABKAAAABRiWFlaAAABPAAAABR3dHB0AAABUAAAABRyVFJDAAABZAAAAChnVFJDAAABZAAAAChiVFJDAAABZAAAAChjcHJ0AAABjAAAADxtbHVjAAAAAAAAAAEAAAAMZW5VUwAAAAgAAAAcAHMAUgBHAEJYWVogAAAAAAAAb6IAADj1AAADkFhZWiAAAAAAAABimQAAt4UAABjaWFlaIAAAAAAAACSgAAAPhAAAts9YWVogAAAAAAAA9tYAAQAAAADTLXBhcmEAAAAAAAQAAAACZmYAAPKnAAANWQAAE9AAAApbAAAAAAAAAABtbHVjAAAAAAAAAAEAAAAMZW5VUwAAACAAAAAcAEcAbwBvAGcAbABlACAASQBuAGMALgAgADIAMAAxADb/2wBDAAUDBAQEAwUEBAQFBQUGBwwIBwcHBw8LCwkMEQ8SEhEPERETFhwXExQaFRERGCEYGh0dHx8fExciJCIeJBweHx7/2wBDAQUFBQcGBw4ICA4eFBEUHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh7/wAARCABGAK0DASIAAhEBAxEB/8QAHQAAAwEAAgMBAAAAAAAAAAAAAAcIBgEFAgQJA//EAEsQAAECBQEBCQoKBwkBAAAAAAECAwAEBQYRByEIEhMYMVFWk9EUFRdBVGFxgZKUFiQyNjdVc3SxsiMzQlKRocMiU2JkcnWCg7Pw/8QAGwEAAQUBAQAAAAAAAAAAAAAABAABAwUGAgf/xAA5EQABAwMBAwkGBAcBAAAAAAABAAIDBAURMQYhQRITFBYyUVKR0VNUYXGBsRWSweEHIiMzNULwwv/aAAwDAQACEQMRAD8AsuPQr1ZpdBpjtTrE6zJSbWN+66cAZ5I863U5OjUmZqdQeSzLSzZW4tRwAB/9iIY1f1GqmoFwOTLzrrVLaURJypOAhPOedRgWpqRCO8labZrZuW9THfyY26n9B8fsnvdO6Yt6Tddl6HS5qecQogPLwGl+cYOYy/Ghq/RyS9pfbE8QRVmrnJ7S9Xg2Js8TeSYuV8SSqH40NX6OSXtL7YONDV+jkl7S+2J4ghulTeJTdTrN7AeZ9VQ/Ghq/RyS9pfbBxoav0ckvaX2xPEELpU3iS6nWb2A8z6qh+NDV+jkl7S+2GloRqjOajd8O66axJ9y43vBknfZ9JiJopfcUfJrv/H8RE9NUSulDXHcs9tTs1bKK1yTQRAOGMHJ7x8VSkLnXHU+W06pUqptlubqU2v8AQy6ycFA+Uo45siNtcdYkaBRJur1F0NS0s2pxZJ2nA5B5zyCII1Ku2cvW8J2uzalb11e9YQf2GhsSMc+MZgusqDE3kt1Kx2x+zou1SZJh/SZr8TwH6n904+NDV+jkl7S+2DjQ1fo5Je0vtieUpUtQShJUonAAG0x5zLD0s+tiYbU06g4UhQ2iK3pU3iXqPU+y5xzI8z6qkrd3SdVqlfkKau35NCZqYQ0VBSsjfKAzy+eHjqXcbtp2bPV1iXRMOSychtecH+EQpp/8+KJ9+a/OIs/dDfRNWfs+2DaaaR0by47wsNtPZKGjuVJDBHhrzvG/fvCTSd1FWCkE25I5x+8vtjnjQ1fo5Je0vtidxyCOYBFXNjtLddTrN7AeZ9VQ/Ghq/RyS9pfbBxoav0ckvaX2xPEEP0qbxJdTrN7AeZ9VQx3UNXxstyS9pfbFBaa3E7dlk064HpdEu5NoKi2nOE4UR4/RHz2PJF3bnr6H6B9ir86oLop5JJCHHO5Yvbew0Fuo45KaPkkuxx0we9YrdjV+Yp1iydHlnQkVOY3kwnxlCRvh/MCJHijt2vLTAnaHNkHudSS2Obf7T+ETjAtYSZ3ZWs2GhZFZ4y3VxJPzzhMvRLSWoaizD8y7Mqp9IlzvXJgJytajnYgHYcY282yHCNy9bgABuOpHz8EiO63IE/Jv6UJkGXkKmpSbdMw2OVG/USnPpA/lDlg2mpYnRBzhklYTaPau6w3KWGKTkNacAADhx3jjqkBxX7c6R1LqkQcV+3OkdS6pEP8AggjocHhVH1vvPtz5D0SA4r9udI6l1SIOK/bnSOpdUiH/AAQuhweFLrfefbnyHokBxX7c6R1LqkRv9I9Lqdpz3b3BUpmd7rxvuGQE73HNiGBC618v9ixbLecaVvqpOgsybfnPKo8wAz64YwwwDnMYwkLxeLy4UJlLuWQMYHpw1SZ3WmoiqlVPgTSpgdxyi8zykKzwjoPyD/pI/nE/x5vuuzD7j77inXXFFS1qOSpR5SY7mxLZn7vumSoFOQsuTCxv1pTng28/2lnzCKaR7pXlx1P/AGF7fbaGnstAIgcNaMk954kpqblXT1dcuA3bUWR3upqvi6Vj9a94jjkKRt9cKe9yVXfVSfKVfjF92nb9Pte2JWh0xlLUvLNb0AftK5So+k5MQJevzuqv3lf4wRUwczGwcd+Vl9lry+73SpnPZAAaO4ZP31K/TT/58UT781+cRZ+6G+ias/Z9sRhp/wDPiiffmvziLP3Q30TVn7PtiSk/syf9wQe2X+Xofn/6ChMcghj6CafSGolxz9MqE8/JolpTh0qaSCSd8E42+mFwOQQ+txX8/az/ALb/AFEwHTtDpGtOi1+0lTLS2uaaF2HAbj9QtnxX7c6R1LqkQcV+3OkdS6pEP+CLrocHhXivW+8+3PkPRIA7l63CMfCOpD/qRDksa3Ze1LWkqBKvuPsyiClLiwApWSTtx6Y7uCO46eOM5aMIGvvlfcGCOplLgDnhr9AlPuo7TeubTdyZlEOOzdKX3S00hOSsfJV/AEn1RFg2jMfSpxCXEKQtIUlQIIPIREn6/aJT9Jnpm5LUlVTNOdWVvSjSMrYJP7KRtKfRyQDXU5J5xv1W62D2iigaaCodjJy0nTfqP1CUNl3bX7OqvfK359Uq+U71QxvkLHnSdhhgjdD6i4GZuUJ+7I7IUbqFtOqadQptxJwpKhgg8xEcZHPFeyVzRhrsL0iqtFBWP5yeJrj3kApu8YfUXyqU93R2QcYfUXyqU93R2QosjngyOeOukS+Iobq3afd2eQTd4w+ovlUp7ujsg4w+ovlUp7ujshRZHPBkc8LpEniKXVu0+7s8gm7xh9RfKpT3dHZGFv8AvSvXxV26lX5lLrrTYabShISlKck8g2Z28sZyCGdI94w4koils1BSSc5BC1ru8DeiNZp3f1bsV2afoaZZL8yAFOuNJWpIHiGRszGTyIMjnjgOLTkFGVFPFUxmKZoc06g6Ju8YfUTyqU93T2QqqlOPVCffnZggvPrK1kDG0x6+RzwR06Rz+0coejtlHREmnjDCdcDC9mlTr9NqUtUJUgPS7iXGyRkZByIYV062XtclEmaPUX5YysyMLCWEg49IELSDIhg9zRgHGV3UW+lqZGyTRhzm6EjT5IjTaeXxXLFqcxUKE40h6YZ4FfCIChvcg+PziMzBDAlpyFNPBHURmKVoc06g6Ju8YfUXyqU93R2QcYfUXyqU93R2QosjnjjI54k6RL4iqrq3afd2eQV3aB3RVbw02lK5WXELnHXnUKKEBIwlZA2CN9Cm3Jf0LU/7zMf+hhsxeU5Lomk9wXgt9iZDcp44xhoc4ADgMogO0YMEETKqWTuTTizK/kz9Bk+EJypxppKFqPnIG2M2dBtOCc96n+vMNCCIjDG45LQrGG8V8LeTHM4D5lK/wDacfVb/AF5g8A2nH1W/15hoQQ3R4vCFN+PXP3h/5ilf4BtOPqt/rzGO1f0/0tsKzpiruU11U4v9FJtF8krcIONnNs2mH5NzDEpKuzUy6lplpBW4tRwEpHKTENa9agLvu9XX2VFFMkiWZNsqzkDYpfrIyPTAlZzULNzRk6LUbKfil3rAHzv5tu938x+g+v2S/Ud8oqAxk5xzR2NsUWeuKvydEprZcmZpwISB4h4z6hkx1m+TziKy3J+nXeaim76tLAT0+jEolYBLTX73OlR2+qK6CIzPDQvTtobyy0UTpj2tGjvP7alaCkbn+wZWly0vOyj81MttpS69wpTwisbVY8WeaPa8A2nH1W/15hoQReCmiH+oXhTtoLo4kmd3mUrzoPpwkE96n9g/vzEbXTLMSVx1CUlklLLT6koSTnAEfRhz9Wr0GPnXeqk/C6qkKB+Mr8fniur2Mj5PJGF6F/D64VVXNOJ5C7AGMkniVxaEoxP3VS5KaSVsPzTbbiQcZSVAGKZ1h0hsa39PqpVqbTXm5plGW1F4kA+iJssBaRfFEyoD48z4/wDGIs/dDkDSWskkAcH4/XHFKxjonkjOPRH7WVtRBdKOOJ5aHHeAcZ3jVQqOQQ2dzFZ9CvO7KnIV+WW+yzI8K2Er3uFb9Iz/AAMKQKTgbRD63FSgb9rIBGe9v9RMDU2HSNBWj2nmkhtM8kbiHAbiNdQnH4BtOPqt/rzB4BtOPqt/rzDQgi96PF4QvCvx65+8P/MV09nW1SbToTdForKmZNtSlpSpW+IKjk7Y7iCCJQABgKrllfK8veck7yTxRBBBDrhEEEEJJEEEEJJfhUZOWqEk7JTjKXpd5JQ42oZC0nlB80ZYaX6eAACzqOAP8sIII5cxru0MoiGrngGInlo+BI+y5GmOnoIIs+jgg5HxcRrW0IbbS22kJQkYSkDYBBBCaxreyMJpqmafHOvLsd5J+68oIII6UCDtGDGTf02sJ99b71pUlxxZ3ylKlwSTzwQRy5jXdoZU0NTNBkxPLc9xI+y5ltN7DlphuYl7TpLbzSgtC0y4BSoHIIjQ1amyFWkXJGpSjU3LOfLadTlKvSIIIQY1owAnkqp5HBz3kkaZJOPks0NMNPQMfA+j+7COyt6z7Xt6bcm6HQpCnvuI4NbjDQSVJznB82YIIYRMByAFI+vqpGlr5HEHgSV3sEEEdoREEEEJJf/Z"

# Écrire le logo une fois au démarrage
_LOGO_PATH = "/tmp/caf_logo_server.png"
try:
    with open(_LOGO_PATH, "wb") as _f:
        _f.write(base64.b64decode(LOGO_B64))
except:
    _LOGO_PATH = None


def generate_xlsx(data: dict) -> bytes:
    wb = Workbook(); ws = wb.active; ws.title = "Constat"

    col_w = {'A':6,'B':6,'C':7,'D':7,'E':7,'F':9,'G':9,'H':9,'I':9,'J':9,'K':9}
    for col, w in col_w.items():
        ws.column_dimensions[col].width = w

    for r,h in {1:22,2:22,3:15,4:14,5:14,6:18,7:22,8:15,9:15,10:20,
                11:15,12:15,13:20,14:15,15:18,16:20,17:20,18:20,19:18}.items():
        ws.row_dimensions[r].height = h

    DG="595959"; BK="000000"; WH="FFFFFF"
    GR="00B050"; RD="FF0000"; PK="FFCCCC"; LG="CCFFCC"
    YL="FFFF99"; GY="D9D9D9"; TL="9CC2E5"; OL="E2EFDA"; RC="FF0000"

    def fnt(bold=False, size=10, color=BK):
        return Font(name="Arial", bold=bold, size=size, color=color)
    def fill(c): return PatternFill("solid", fgColor=c)
    def aln(h="center", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    def brd():
        s = Side(style="thin", color=BK)
        return Border(left=s, right=s, top=s, bottom=s)

    def M(r1, c1, r2, c2, val=None, bold=False, size=10, color=BK,
          bg=None, ha="center", wrap=False):
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
        cell = ws.cell(row=r1, column=c1, value=val)
        cell.font = fnt(bold=bold, size=size, color=color)
        cell.alignment = aln(ha, "center", wrap)
        if bg: cell.fill = fill(bg)
        cell.border = brd()
        return cell

    # R1-2 Logo + header
    ws.merge_cells('A1:B2')
    ws['A1'].fill = fill(WH); ws['A1'].border = brd()
    if _LOGO_PATH and os.path.exists(_LOGO_PATH):
        try:
            img = XLImage(_LOGO_PATH)
            img.width = 72; img.height = 38; img.anchor = 'A1'
            ws.add_image(img)
        except:
            ws['A1'].value = "CAF"
            ws['A1'].font = fnt(bold=True, size=14, color=RC)
    else:
        ws['A1'].value = "CAF"
        ws['A1'].font = fnt(bold=True, size=14, color=RC)

    M(1,3,2,7,"VERIFICATION\nETALON SECONDAIRE DE PEINTURE",
      bold=True,size=12,color=WH,bg=DG,wrap=True)
    M(1,8,1,9,"Constat N°",bold=True,size=10,color=WH,bg=DG)
    M(1,10,1,11,"Site de "+data.get('site','Reichshoffen'),bold=True,size=9,color=WH,bg=DG)
    M(2,8,2,11,data['numero_constat'],bold=True,size=10,color=BK,bg=YL)

    # R3-5 Référence étalon
    M(3,1,3,5,"Référence de l'étalon:",size=9,bg=GY,ha="left")
    M(3,6,3,11,"Teinte : "+str(data['teinte']),size=9,bg=WH,ha="left")
    M(4,1,4,5,"",bg=GY)
    M(4,6,4,11,"Brillance : "+str(data['brillance']),size=9,bg=WH,ha="left")
    M(5,1,5,5,"",bg=GY)
    M(5,6,5,11,"Technologie : "+str(data['technologie']),size=9,bg=WH,ha="left")

    # R6-8 Spectro
    M(6,1,6,4,"Référence spectro-colorimètre",size=9,bg=GY,ha="left")
    M(6,5,6,5,data['spectro_ref'],size=9,bg=WH)
    M(6,6,6,11,"N° ETALON SECONDAIRE A VALIDER",bold=True,size=10,color=WH,bg=DG)
    M(7,1,7,4,"N° d'identification",size=9,bg=GY,ha="left")
    M(7,5,7,5,data['spectro_id'],size=9,bg=WH)
    M(7,6,7,11,data['etalon_secondaire'],bold=True,size=14,color=BK,bg=WH)
    M(8,1,8,4,"Validité",size=9,bg=GY,ha="left")
    M(8,5,8,5,data['spectro_validite'],size=9,color=RC,bg=WH)
    M(8,6,8,11,"Valeur L * a * b",bold=True,size=10,color=WH,bg=DG)

    # R9-10 Primaire
    M(9,1,9,5,"N° ETALON PRIMAIRE DE REFERENCE",bold=True,size=9,bg=GY)
    M(9,6,9,7,"L",bold=True,size=11,bg=OL)
    M(9,8,9,9,"a",bold=True,size=11,bg=OL)
    M(9,10,9,11,"b",bold=True,size=11,bg=OL)
    M(10,1,10,5,data['etalon_primaire'],bold=True,size=12,bg=WH)
    M(10,6,10,7,data['ref_L'],bold=True,size=12,bg=YL)
    M(10,8,10,9,data['ref_a'],bold=True,size=12,bg=YL)
    M(10,10,10,11,data['ref_b'],bold=True,size=12,bg=YL)

    # R11-13 Deltas
    M(11,1,11,3,"Date de validation",size=9,bg=GY,ha="left")
    M(11,4,11,5,data['date_validation_primaire'],size=9,bg=WH)
    M(11,6,11,7,"ΔL",bold=True,size=11,bg=OL)
    M(11,8,11,9,"Δa",bold=True,size=11,bg=OL)
    M(11,10,11,11,"Δb",bold=True,size=11,bg=OL)

    tL = f"{data['tol_L_min']}  |  {data['tol_L_max']}"
    ta = f"{data['tol_a_min']}  |  {data['tol_a_max']}"
    tb = f"{data['tol_b_min']}  |  {data['tol_b_max']}"
    M(12,1,12,5,"",bg=WH)
    M(12,6,12,7,tL,size=8,color=WH,bg=BK)
    M(12,8,12,9,ta,size=8,color=WH,bg=BK)
    M(12,10,12,11,tb,size=8,color=WH,bg=BK)

    dL = round(data['mes_L'] - data['ref_L'], 3)
    da = round(data['mes_a'] - data['ref_a'], 3)
    db = round(data['mes_b'] - data['ref_b'], 3)
    dE = round(math.sqrt(dL**2 + da**2 + db**2), 3)
    okL = data['tol_L_min'] <= dL <= data['tol_L_max']
    oka = data['tol_a_min'] <= da <= data['tol_a_max']
    okb = data['tol_b_min'] <= db <= data['tol_b_max']
    okE = dE <= data['tol_dE']
    ok  = okL and oka and okb and okE
    dbg = LG if ok else PK

    M(13,1,13,5,"",bg=WH)
    M(13,6,13,7,dL,bold=True,size=12,bg=dbg)
    M(13,8,13,9,da,bold=True,size=12,bg=dbg)
    M(13,10,13,11,db,bold=True,size=12,bg=dbg)

    # R14-16 Mesures
    M(14,1,14,5,"Valeur L * a * b",bold=True,size=10,bg=GY)
    M(14,6,14,11,"ΔE",bold=True,size=12,bg=OL)
    M(15,1,15,2,"L",bold=True,size=11,bg=TL)
    M(15,3,15,4,"a",bold=True,size=11,bg=TL)
    M(15,5,15,6,"b",bold=True,size=11,bg=TL)
    M(15,7,15,11,f"≤ {data['tol_dE']}",size=9,color=WH,bg=BK)
    M(16,1,16,2,data['mes_L'],bold=True,size=12,bg=YL)
    M(16,3,16,4,data['mes_a'],bold=True,size=12,bg=YL)
    M(16,5,16,6,data['mes_b'],bold=True,size=12,bg=YL)
    M(16,7,16,11,dE,bold=True,size=13,bg=LG if okE else PK)

    # R17-18 Résultats
    M(17,1,17,5,"Résultat du contrôle colorimétrique:",bold=True,size=9,bg=GY,ha="left")
    M(17,6,17,11,"CONFORME" if ok else "NON CONFORME",bold=True,size=13,color=WH,bg=GR if ok else RD)
    rv   = data.get('res_visuel','CONFORME')
    norme = data.get('norme_visuelle','DTREI 150613')
    M(18,1,18,5,f"Résultat du contrôle visuel suivant {norme} :",bold=True,size=9,bg=GY,ha="left",wrap=True)
    M(18,6,18,11,rv,bold=True,size=13,color=WH,bg=GR if rv=="CONFORME" else RD)

    # R19 Signataire
    M(19,1,19,3,"Etabli par",bold=True,size=10,bg=GY)
    M(19,4,19,7,data['etabli_par'],bold=True,size=11,bg=WH)
    M(19,8,19,8,"Le",bold=True,size=10,bg=GY)
    M(19,9,19,11,data['date_constat'],bold=True,size=11,bg=WH)

    ws.print_area = 'A1:K19'
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_setup.orientation = 'landscape'

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()


@app.route("/health", methods=["GET"])
def health():
    return {"status": "ok", "message": "Constat Generator API is running"}


@app.route("/generate", methods=["POST"])
def generate():
    data = request.get_json()
    if not data:
        return {"error": "JSON requis"}, 400
    required = ["numero_constat","teinte","brillance","technologie",
                "spectro_ref","spectro_id","spectro_validite",
                "etalon_secondaire","etalon_primaire",
                "ref_L","ref_a","ref_b","date_validation_primaire",
                "tol_L_min","tol_L_max","tol_a_min","tol_a_max",
                "tol_b_min","tol_b_max","tol_dE",
                "mes_L","mes_a","mes_b","etabli_par","date_constat"]
    missing = [f for f in required if f not in data]
    if missing:
        return {"error": f"Champs manquants : {missing}"}, 400
    try:
        xlsx = generate_xlsx(data)
        fname = data['numero_constat'].replace("/","_").replace("°","").replace(" ","_") + ".xlsx"
        return send_file(io.BytesIO(xlsx),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True, download_name=fname)
    except Exception as e:
        return {"error": str(e)}, 500


if __name__ == "__main__":
    print("CAF Constat Generator — http://localhost:5000")
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5000)), debug=True)
